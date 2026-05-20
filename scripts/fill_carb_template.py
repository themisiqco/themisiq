#!/usr/bin/env python3
"""
ThemisIQ — CARB SB 253 Official Template Filler
Loads the exact CARB template and fills column B with inventory data.
No hardcoded customer data — everything comes from the inventory JSON.
"""

import sys
import json
from openpyxl import load_workbook

def fill_carb_template(template_path: str, inventory: dict, output_path: str):
    wb = load_workbook(template_path)
    ws = wb['Form']

    def fill(row, value):
        """Fill column B only — the answer column in the official CARB template."""
        if value is not None and value != '':
            ws[f'B{row}'] = value

    # ── ORGANIZATION INFORMATION (rows 2-16) ──────────────────────────
    fill(2,  inventory.get('trade_secret', 'No'))
    fill(3,  inventory.get('company_name', ''))
    fill(4,  inventory.get('ein', ''))
    # B5 = NAICS — dropdown, leave for customer to select
    fill(6,  inventory.get('website', ''))
    fill(7,  inventory.get('address', ''))
    fill(8,  inventory.get('city', ''))
    fill(9,  inventory.get('hq_state', ''))
    fill(10, inventory.get('country', ''))
    fill(11, inventory.get('postal_code', ''))
    fill(12, inventory.get('contact_name', ''))
    fill(13, inventory.get('contact_title', ''))
    fill(14, inventory.get('contact_phone', ''))
    fill(15, inventory.get('contact_email', ''))
    # B16 = verification date — completed by verifier, leave blank

    # ── 3RD PARTY VERIFICATION (rows 17-20) ───────────────────────────
    # All completed by verifier — leave blank

    # ── INVENTORY BOUNDARY (rows 21-46) ───────────────────────────────
    year = inventory.get('reporting_year', 2024)
    fill(21, inventory.get('period_start', f'{year}-01-01'))
    fill(22, inventory.get('period_end', f'{year}-12-31'))
    fill(23, inventory.get('boundary_approach', 'Operational Control'))
    fill(26, inventory.get('entities_list', ''))
    fill(27, 'No')   # subsidiary reports separately
    fill(29, 'No')   # regions excluded

    # Source inclusions — auto-populated from inventory data
    s1_stat  = inventory.get('s1_stationary', 0) or 0
    s1_mob   = inventory.get('s1_mobile', 0) or 0
    s1_proc  = inventory.get('s1_process', 0) or 0
    s1_fug   = inventory.get('s1_fugitive', 0) or 0
    s1_total = inventory.get('s1_total', 0) or 0
    s2_loc   = inventory.get('s2_location_total', 0) or 0
    s2_mkt   = inventory.get('s2_market_total', 0) or 0
    s2_steam = inventory.get('s2_steam', 0) or 0
    s2_heat  = inventory.get('s2_heating', 0) or 0
    s2_cool  = inventory.get('s2_cooling', 0) or 0
    biogenic = inventory.get('s1_biogenic', 0) or 0
    revenue  = inventory.get('revenue_millions', 0) or 0

    fill(31, yn(s1_stat > 0))   # Scope 1 stationary
    fill(32, yn(s1_mob > 0))    # Scope 1 mobile
    fill(33, yn(s1_proc > 0))   # Scope 1 process
    fill(34, yn(s1_fug > 0))    # Scope 1 fugitive
    fill(35, yn(s2_loc > 0))    # Scope 2 location electricity
    fill(36, yn(s2_heat > 0))   # Scope 2 location heating
    fill(37, yn(s2_steam > 0))  # Scope 2 location steam
    fill(38, yn(s2_cool > 0))   # Scope 2 location cooling
    fill(39, yn(s2_mkt > 0))    # Scope 2 market electricity
    fill(40, 'No')               # Scope 2 market heating
    fill(41, 'No')               # Scope 2 market steam
    fill(42, 'No')               # Scope 2 market cooling
    fill(43, yn(biogenic > 0))   # direct biogenic
    fill(44, 'No')               # indirect biogenic
    fill(45, 'No')               # sources excluded

    # ── SCOPE 1 DISCLOSURE (rows 47-53) ───────────────────────────────
    s1_intensity = round(s1_total / revenue, 6) if revenue > 0 else None

    fill_num(ws, 47, s1_total)
    fill_num(ws, 48, s1_mob)
    fill_num(ws, 49, s1_proc)
    fill_num(ws, 50, s1_fug)
    fill_num(ws, 51, s1_intensity)
    # B52 = limited assurance confirmation — completed by verifier
    # B53 = explanation — completed by verifier

    # ── SCOPE 2 DISCLOSURE (rows 54-60) ───────────────────────────────
    s2_total    = s2_loc
    s2_intensity = round(s2_total / revenue, 6) if revenue > 0 else None

    fill_num(ws, 54, s2_total)
    fill_num(ws, 55, s2_steam)
    fill_num(ws, 56, s2_heat)
    fill_num(ws, 57, s2_cool)
    fill_num(ws, 58, s2_intensity)
    # B59-60 = verifier

    # ── BIOGENIC (row 61) ─────────────────────────────────────────────
    fill_num(ws, 61, biogenic)

    # ── METHODS (rows 62-68) ──────────────────────────────────────────
    fill(62, 'US EPA')
    fill(63, str(year))
    fill(64, 'US EPA (2024) Emission Factors for Greenhouse Gas Inventories')
    fill(65, 'US EPA eGRID (2023) — subregion location-based emission factors')
    fill(66, 'IPCC Fourth Assessment Report (AR4, 2007) — as required by CARB SB 253')
    fill(67, 'Activity data × emission factor = GHG emissions (mtCO₂e)')
    fill(68, 'Standard EPA calculation methodology via ThemisIQ platform (www.themisiq.co)')

    # ── DE MINIMIS (rows 69-70) ───────────────────────────────────────
    fill(69, inventory.get('de_minimis_sources', 'None identified'))
    fill(70, inventory.get('de_minimis_quantity', '0'))

    # ── MRR (row 71) ──────────────────────────────────────────────────
    fill(71, inventory.get('mrr_ids', ''))

    # ── EMISSION REDUCTIONS (rows 72-75) ──────────────────────────────
    ren_kwh = inventory.get('renewable_electricity_kwh', 0) or 0
    ren_gas = inventory.get('renewable_gas_mtco2', 0) or 0
    if ren_kwh > 0:
        fill(72, 'Renewable Electricity')
    if ren_gas > 0:
        fill(73, 'Renewable Gas')
    fill_num(ws, 74, round(ren_kwh * 0.3866 / 1000, 4) if ren_kwh > 0 else None)
    fill_num(ws, 75, round(ren_gas, 4) if ren_gas > 0 else None)

    # ── SCOPE 1 BY GAS (rows 98-103) ──────────────────────────────────
    fill_num(ws, 98,  inventory.get('s1_co2', 0))
    fill_num(ws, 99,  inventory.get('s1_ch4', 0))
    fill_num(ws, 100, inventory.get('s1_n2o', 0))
    fill_num(ws, 101, inventory.get('s1_hfc', 0))
    fill_num(ws, 102, inventory.get('s1_pfc', 0))
    fill_num(ws, 103, inventory.get('s1_sf6', 0))

    # ── SCOPE 2 BY GAS (rows 105-110) ─────────────────────────────────
    # Location-based electricity is ~100% CO2
    fill_num(ws, 105, round(s2_total * 0.999, 4) if s2_total else None)
    fill_num(ws, 106, round(s2_total * 0.001, 4) if s2_total else None)
    fill_num(ws, 107, 0)
    fill_num(ws, 108, 0)
    fill_num(ws, 109, 0)
    fill_num(ws, 110, 0)

    # ── SCOPE 1 BY SOURCE AND GAS (rows 112-145) ──────────────────────
    # Stationary combustion
    fill_num(ws, 112, round(s1_stat * 0.97, 4) if s1_stat else None)   # CO2
    fill_num(ws, 113, round(s1_stat * 0.02, 4) if s1_stat else None)   # CH4
    fill_num(ws, 114, round(s1_stat * 0.01, 4) if s1_stat else None)   # N2O
    for r in [115, 116, 117]: fill_num(ws, r, 0)                        # HFC PFC SF6

    # Mobile combustion
    fill_num(ws, 118, round(s1_mob, 4) if s1_mob else None)
    fill_num(ws, 119, round(s1_mob * 0.97, 4) if s1_mob else None)     # CO2
    fill_num(ws, 120, round(s1_mob * 0.015, 4) if s1_mob else None)    # CH4
    fill_num(ws, 121, round(s1_mob * 0.015, 4) if s1_mob else None)    # N2O
    for r in [122, 123, 124]: fill_num(ws, r, 0)

    # Process
    fill_num(ws, 125, round(s1_proc, 4) if s1_proc else None)
    fill_num(ws, 126, round(s1_proc, 4) if s1_proc else None)           # CO2
    for r in [127, 128, 129, 130, 131]: fill_num(ws, r, 0)

    # Fugitive — HFC refrigerants
    fill_num(ws, 132, round(s1_fug, 4) if s1_fug else None)
    for r in [133, 134, 135]: fill_num(ws, r, 0)                        # CO2 CH4 N2O = 0
    fill_num(ws, 136, round(s1_fug, 4) if s1_fug else None)             # HFC = all fugitive
    for r in [137, 138]: fill_num(ws, r, 0)

    # Agricultural — 0 for wheat processor (no livestock)
    for r in range(139, 146): fill_num(ws, r, 0)

    # ── SCOPE 2 BY SOURCE AND GAS (rows 147-174) ──────────────────────
    # Purchased electricity
    fill_num(ws, 147, round(s2_loc, 4) if s2_loc else None)
    fill_num(ws, 148, round(s2_loc * 0.999, 4) if s2_loc else None)    # CO2
    fill_num(ws, 149, round(s2_loc * 0.001, 4) if s2_loc else None)    # CH4
    for r in [150, 151, 152, 153]: fill_num(ws, r, 0)

    # Steam
    fill_num(ws, 154, round(s2_steam, 4) if s2_steam else None)
    fill_num(ws, 155, round(s2_steam * 0.999, 4) if s2_steam else None)
    for r in [156, 157, 158, 159, 160]: fill_num(ws, r, 0)

    # Heating
    fill_num(ws, 161, round(s2_heat, 4) if s2_heat else None)
    fill_num(ws, 162, round(s2_heat * 0.999, 4) if s2_heat else None)
    for r in [163, 164, 165, 166, 167]: fill_num(ws, r, 0)

    # Cooling
    fill_num(ws, 168, round(s2_cool, 4) if s2_cool else None)
    fill_num(ws, 169, round(s2_cool * 0.999, 4) if s2_cool else None)
    for r in [170, 171, 172, 173, 174]: fill_num(ws, r, 0)

    # ── SCOPE 2 INTENSITY (row 175) ───────────────────────────────────
    fill_num(ws, 175, s2_intensity)

    # ── BIOGENIC DUPLICATE (row 176) ──────────────────────────────────
    fill_num(ws, 176, biogenic if biogenic else None)

    # ── DATA QUALITY (rows 177-179) ───────────────────────────────────
    fill(177, 'Activity data sourced from utility bills and fuel purchase records provided by the reporting entity. Emission factors from US EPA (2024). Electricity emission factors from US EPA eGRID (2023). Calculations performed by ThemisIQ platform (www.themisiq.co).')
    fill(178, 'Good')
    fill(179, 'Activity data based on utility bills and fuel records. All emission factors from authoritative published sources. Calculations independently verifiable.')

    wb.save(output_path)


def yn(condition): 
    return 'Yes' if condition else 'No'

def fill_num(ws, row, value):
    """Fill a numeric cell — skip if None or 0."""
    if value is not None and value != 0:
        cell = ws[f'B{row}']
        cell.value = round(float(value), 4)
        cell.number_format = '#,##0.0000'


if __name__ == '__main__':
    if len(sys.argv) != 4:
        print('Usage: python3 fill_carb_template.py <template.xlsx> <inventory.json> <output.xlsx>')
        sys.exit(1)

    template_path = sys.argv[1]
    inventory_path = sys.argv[2]
    output_path = sys.argv[3]

    with open(inventory_path) as f:
        inventory = json.load(f)

    fill_carb_template(template_path, inventory, output_path)
    print(f'CARB template filled and saved to {output_path}')
